import openpyxl
from openpyxl.styles import PatternFill

lesson_colors = {'digital':'ea2c49','ds':'308d46','elec':'0b54e5','sport':'9b1700','az-digital':'e23b1d','curcuit':'03c13c','engi-prob':'eabb2e','engi-math':'dee509','eng':'09dee5'}

lesson_patterns = {}

print('start')

excel_path = '_primary.xlsx'

wb = openpyxl.load_workbook(excel_path)
work_sheet = wb["primary"]


for lesson, color in lesson_colors.items():
    lesson_patterns[lesson] = PatternFill(patternType='solid', fgColor= color)


for row in work_sheet.rows:
    for cell in row:
        if(cell.value):
            s = cell.value.split()
            if(s[0] in lesson_patterns):
                cell.fill = lesson_patterns[s[0]]



wb.save("new_excel.xlsx")

print('done')
