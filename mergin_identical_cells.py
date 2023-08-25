from openpyxl import load_workbook

# target_excel_file = "new_format_excel.csv"
# sheet_name = 'new_format_excel'
target_excel_file = 'formated_excel.xlsx'
sheet_name = 'new_format_excel'
workbook = load_workbook(filename=target_excel_file)

list_of_merge = []

last_value = None

# current_row = 0
# current_column = 0


ws = workbook[sheet_name]

row_count = ws.max_row
column_count = ws.max_column

start = 1
count = 0

for row in range(1, row_count + 1):
    last_value = ws.cell(row= row,column=1).value
    for column in range(2, column_count + 1):
        if(last_value == ws.cell(row= row, column=column).value):
            count += 1

        else:
            if(count != 0):
                list_of_merge.append((row,start,count))
                count = 0
            start = column
            last_value = ws.cell(row=row, column = column).value

    if(count != 0):
        list_of_merge.append((row,start,count))
    start = 1
    count = 0

for item in list_of_merge:
    row, start_column, count = item
    ws.merge_cells(start_row=row,end_row=row, start_column=start_column, end_column= start_column + count)



workbook.save(filename=target_excel_file)

        

#ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=2)
#workbook.save(filename=target_excel_file)
